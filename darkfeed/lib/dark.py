from datetime import datetime as dt
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font

class DarkFeed:
    def __init__(self) -> None:

        self.latam = ["Argentina", "Bolivia", "Brazil", "Chile", "Colombia", "Ecuador", "Guyana", "Paraguay", "Peru", "Suriname", "Uruguay", "Venezuela", "Guatemala", "El Salvador", "Honduras", "Nicaragua", "Costa Rica", "Panama", "Belize", "Mexico", "Bahamas", "Haiti", "Jamaica", "Puerto Rico", "Trindad and Tobago"]
        self.south_america = ["Argentina", "Bolivia", "Brazil", "Chile", "Colombia", "Ecuador", "Guyana", "Paraguay", "Peru", "Suriname", "Uruguay", "Venezuela", "Bahamas", "Cuba", "Falkland"]
        self.central_america = ["Guatemala", "El Salvador", "Honduras", "Nicaragua", "Costa Rica", "Panama", "Belize", "Bahamas", "Haiti", "Jamaica", "Puerto Rico", "Trindad and Tobago"]
        self.north_america = ["United States", "Canada", "Mexico"]
        self.europe = ["Albania", "Andorra", "Austria", "Belarus", "Belgium", "Bosnia", "Bulgaria", "Croatia", "Czech Republic", "Denmark", "Estonia", "Finland", "France", "Germany", "Greece", "Hungary", "Iceland", "Ireland", "Italy", "Latvia", "Liechtenstein", "Lithuania", "Luxembourg", "Malta", "Moldova", "Monaco", "Montenegro", "Netherlands", "Macedonia", "Norway", "Poland", "Portugal", "Romania", "San Marino", "Serbia", "Slovakia", "Slovenia", "Spain", "Sweden", "Switzerland", "Ukraine", "United Kingdom", "Vatican", "Kosovo", "Iceland", "Russia"]
        self.asia = ["Armenia", "Azerbaijan", "Bangladesh", "Bhutan", "Brunei", "Cambodia", "China", "Georgia", "India", "Indonesia", "Japan", "Kyrgyzstan", "Laos", "Malaysia", "Maldives", "Mongolia", "Myanmar", "Nepal", "North Korea", "Philippines", "Singapore", "South Korea", "Sri Lanka", "Taiwan", "Tajikistan", "Thailand", "Timor-Leste", "Turkmenistan", "Uzbekistan", "Vietnam", "Yemen", "Cyprus", "Northern Cyprus", "Afghanistan", "United Arab Emirates", "Iran", "Iraq", "Israel", "Jordan", "Kazakhstan", "Korea", "Kuwait", "Lebanon", "Oman", "Pakistan", "Palestine", "Qatar", "Syria", "Turkey"]
        self.africa = ["Algeria", "Angola", "Benin", "Botswana", "Burkina Faso", "Burundi", "Cabo Verde", "Cameroon", "Central African Republic", "Chad", "Comoros", "Congo", "Djibouti", "Egypt", "Equatorial Guinea", "Eritrea", "Eswatini", "Ethiopia", "Gabon", "Gambia", "Ghana", "Guinea", "Guinea-Bissau", "Kenya", "Lesotho", "Liberia", "Libya", "Madagascar", "Malawi", "Mali", "Mauritania", "Mauritius", "Morocco", "Mozambique", "Namibia", "Niger", "Nigeria", "Rwanda", "Sao Tome and Principe", "Senegal", "Seychelles", "Sierra Leone", "Somalia", "South Africa", "South Sudan", "Sudan", "Tanzania", "Togo", "Tunisia", "Uganda", "Zambia", "Zimbabwe", "Ivory Coast", "Côte d'Ivoire", ]
        self.oceania = ["Australia", "Fiji", "Kiribati", "Marshall Islands", "Micronesia", "Nauru", "New Zealand", "Palau", "Papua New Guinea", "Samoa", "Solomon Islands", "Tonga", "Tuvalu", "Vanuatu", "New Caledonia"]

        self.sector_manufacturing = ['Manufacturing', 'ManufacturingAndAutomotive']
        self.sector_finance = ['Financial', 'Banking', 'Insurance']
        self.sector_energy = ['Energy']
        self.sector_health = ['HealthCare']

    def count_countries(self, data:list) -> dict:
        """
        A counter victims by country
        """
        count = {}
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                count.setdefault(item.get('Country'), 0)
                count[item.get('Country')] += 1
        return count

    def count_group_name(self, data:list) -> dict:
        """
        A counter victims by ransomware group
        """
        count = {}
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                ransomware = item.get('Group Name').lower()
                count.setdefault(ransomware, 0)
                count[ransomware] += 1
        return count

    def count_sector(self, data:list) -> dict:
        """
        A counter victims by sector
        """
        count = {}
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                if item.get('Sector') != 'Analyzing Victim Data':
                    count.setdefault(item.get('Sector'), 0)
                    count[item.get('Sector')] += 1
        return count

    def filter_country(self, countries:list, data:list) -> list:
        """
        A filter of data based on a country or a list of countries
        """
        selected_items = []
        
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                if item.get('Country') in countries:
                    selected_items.append(item)
        return selected_items

    def filter_sector(self, sectors:list, data:list) -> list:
        """
        A filter of data based on a sector or a list of sectors
        """
        selected_items = []
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                if item.get('Sector') in sectors:
                    selected_items.append(item)
        return selected_items

    def filter_ransomware(self, ransomwares:list, data:list) -> list:
        """
        A filter of data based on a ransomware or a list of ransomwares
        """
        selected_items = []
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                if item.get('Group Name') in ransomwares:
                    selected_items.append(item)
        return selected_items

    def filter_company(self, company_name:list, data:list):
        """
        A searcher for companies based substring
        """
        selected_items = []
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                if company_name.casefold() in item.get('Company').casefold():
                    selected_items.append(item)
        return selected_items

    def filter_after(self, date:str, data:list) -> list:
        """
        A filter of data, for get data after a date
        """
        date_dt = dt.strptime(date, "%Y-%m-%dT%H:%M:%S")
        selected_items = []
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                if dt.strptime(item["Date"], "%Y-%m-%dT%H:%M:%S") >=  date_dt:
                    selected_items.append(item)
        return selected_items

    def filter_before(self, date:str, data:list) -> list:
        """
        A filter of data, for get data before a date
        """
        date_dt = dt.strptime(date, "%Y-%m-%dT%H:%M:%S")
        selected_items = []
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                if dt.strptime(item["Date"], "%Y-%m-%dT%H:%M:%S") <=  date_dt:
                    selected_items.append(item)
        return selected_items

    def get_top_x_countries(self, num:int, data:list) -> dict:
        count = {}
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                if item.get('Country') != 'Analyzing Victim Data':
                    count.setdefault(item.get('Country'), 0)
                    count[item.get('Country')] += 1
        
        sorted_count = dict(sorted(count.items(), key=lambda item: item[1], reverse=True))
        x_items = {k: sorted_count[k] for k in list(sorted_count.keys())[:num]}
        return x_items

    def get_top_x_sectors(self, num:int, data:list) -> dict:
        count = {}
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                if item.get('Sector') != 'Analyzing Victim Data':
                    count.setdefault(item.get('Sector'), 0)
                    count[item.get('Sector')] += 1
        
        sorted_count = dict(sorted(count.items(), key=lambda item: item[1], reverse=True))
        x_items = {k: sorted_count[k] for k in list(sorted_count.keys())[:num]}
        return x_items

    def get_top_x_ransomwares(self, num:int, data:list) -> dict:
        count = {}
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                ransomware = item.get('Group Name').lower()
                count.setdefault(ransomware, 0)
                count[ransomware] += 1
        
        sorted_count = dict(sorted(count.items(), key=lambda item: item[1], reverse=True))
        x_items = {k: sorted_count[k] for k in list(sorted_count.keys())[:num]}
        return x_items
    
    def get_country_list(self, data:list) -> list:
        """
        Get a list of countries that have at least one victim of ransomware
        """
        return list(self.count_countries(data).keys())

    def get_sector_list(self, data:list) -> list:
        """
        List all sectos identified by the victims of ransomware
        """
        return list(self.count_sector(data).keys())

    def get_ransomware_list(self, data) -> list:
        """
        List all ransomwares in the base
        """
        return list(self.count_group_name(data).keys())
    
    def get_cyber_news(self, data) -> list:
        """
        List only events related to CTI info
        """
        selected_items = []
        for item in data:
            if(item.get('Category') != 'Ransomware'):
                selected_items.append(item)
        return selected_items
    
    def get_ransomware_news(self, data:list) -> list:
        """
        List only events related to ransowmare compromise
        """
        selected_items = []
        for item in data:
            if(item.get('Category') == 'Ransomware'):
                selected_items.append(item)
        return selected_items
    
    def count_data_per_month(self, data:list) -> dict:
        """
        Split the data by month and return it in list format
        """
        count = {}
        for item in data:
            date_str = item.get('Date')
            date_dt = dt.strptime(date_str, '%Y-%m-%dT%H:%M:%S')
            month_year = date_dt.strftime('%Y-%m')
            count.setdefault(month_year, 0)
            count[month_year] += 1    
        
        new_dict = {}
        for key in count:
            # Assuming key is in format '%Y-%m', extract month part \o/
            month = key.split('-')[1]
            new_key = f"{month}"
            new_dict[new_key] = count[key]
        return new_dict
    
    def data_2_csv(self, data):
        filename = 'darkfeed.xlsx'
        
        if data:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "DarkFeed"

            if isinstance(data, list):
                if isinstance(data[0], str):
                    sheet['A1'] = "data"
                    sheet.font = Font(bold=True)
                    for index, item in enumerate(data, start=2):
                        sheet.cell(row=index, column=1).value = item
                elif isinstance(data[0], dict):
                    headers = list(data[0].keys())
                    for col_idx, header in enumerate(headers, start=1):
                        sheet.cell(row=1, column=col_idx).value = header
                        sheet.cell(row=1, column=col_idx).font = Font(bold=True)

                    for row_idx, data_dict in enumerate(data, start=2):
                        for col_idx, header in enumerate(headers, start=1):
                            item = data_dict.get(header, "")
                            if isinstance(item,list):
                                item = str(item)
                            sheet.cell(row=row_idx, column=col_idx).value = item
                             
            elif isinstance(data,dict):
                sheet['A1'] = "data"
                sheet['B1'] = "value"
                row_idx = 2
                for key, value in data.items():
                    sheet.cell(row=row_idx, column=1).value = key
                    sheet.cell(row=row_idx, column=2).value = value
                    row_idx += 1

            else:
                print("This isn't the kind of data to be exported to spreadsheet")
            
            sheet.freeze_panes = 'A2' # Freeze the rows above A2
            wb.save(filename)

        else:
            print("There isn't data")
        
        print(f"File '{filename}' created successfully.")



        
